# ####todo
# implement file name selection
# for the lines selecing specific column, make it not hard coded. ie make it so we can choose whatever column based on the selected paramterers
# removes the DUT info currently, so actually try and fix this


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)


INPUTFILE1 = 'input1.xlsx'
OUTPUTFILE1 = 'parsed1.xlsx'

INPUTFILE2 = 'input2.xlsx'
OUTPUTFILE2 = 'parsed2.xlsx'

COMBINEDFILE = 'combined.xlsx'


columnTrans = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC')
#create the tuple of attributes to look for
keyParametersTX = ('ParameterName', 'channel', 'txChainMask', 'rate', 'powerLevel', 'NumValue')
keyParametersTXHeader = ('ParameterName', 'channel', 'txChainMask', 'rate', 'target power', 'b1 measured power', 'b1 EVM', 'b2 measured power', 'b2 EVM')
keyParametersRX = ('ParameterName', 'channel', 'txChainMask', 'rate', 'powerLevel', 'NumValue')

keyParametersTXrow = ('evm', 'avgTxPower')

def colNumToColStr(numberOfCol):
	return columnTrans[numberOfCol + 1]

def indexNumToColStr(index):
	return columnTrans[index]

##############################parses raw data file and saves it#####################
def parseFileTX(rawFileName, newSaveFileName):

	#load the file
	wb = load_workbook(rawFileName)
	ws = wb.active
	#remove the first 3 rows, which includes the DUT info, so maybe change this later
	ws.delete_rows(0,3)


	#find the column index of all the key parameters and creates a tuple with them
	headerRow = ws[1]
	keyParameterIndexTup = ()
	for i in range(len(keyParametersTX)):
		for x in range(len(headerRow)):
			if (keyParametersTX[i] == headerRow[x].value):
				dog = (x,)
				keyParameterIndexTup = keyParameterIndexTup + dog
				
	#go through and delete the columns that are not in the key parameters			
	i = len(keyParameterIndexTup) - 1
	for z in range( len(headerRow) -1, -1, -1):
		#keyParameterIndexTup[i]
		if (keyParameterIndexTup[i] != z):
			ws.delete_cols(z + 1, 1)
		else:
			i = i - 1


		
	###################remove the rows not of interest#########################


	#Create a list of the testnames
	parameterNameCol = ws[indexNumToColStr(keyParametersTX.index('ParameterName'))]
	parameterNameColList = []
	for m in range(len(parameterNameCol)):
		parameterNameColList.append(parameterNameCol[m].value)
	
	#create a list of the index of the rows of the test names	
	keyParameterIndexCol = []
	for i in range(len(keyParametersTXrow)):
		for x in range(len(parameterNameColList)):
			if (keyParametersTXrow[i] == parameterNameColList[x]):
				keyParameterIndexCol.append(x)

	keyParameterIndexCol.sort()
	#print (keyParameterIndexCol)	
	
	#go through and delete the columns that are not in the key parameters of tests (in backwards order)
	i = len(keyParameterIndexCol) - 1
	for z in range( len(parameterNameCol) -1, -1, -1):
		#saves the column names (if you want to remove it, remove the z !=0 condition)
		if (keyParameterIndexCol[i] != z):
		#if ((keyParameterIndexCol[i] != z) and (z !=0)):
			ws.delete_rows(z + 1, 1)
		else:
			i = i - 1
		
		
	################### put the avgtxpower in same row as the evm ##############################
	#Create a list of the numValues
	numValueColTup = ws[indexNumToColStr(keyParametersTX.index('NumValue'))]
	evmColList = []
	avgTXPowerColList = []
	for m in range(len(numValueColTup)):
		#if we are on an even index, then it's an EVM value, so seperate that out
		if (m%2 == 0):
			evmColList.append(numValueColTup[m].value)
		else:
			avgTXPowerColList.append(numValueColTup[m].value)
		

	#insert a new column before the evm measuremenr
	ws.insert_cols(keyParametersTX.index('NumValue') + 1)
		
	#delete the avgtxpower rows, and add the avgtxpower values to the evm rows from the backwards direction
	for z in range(len(numValueColTup) -1, -1, -1):
		#if on an odd index (i.e. we are on a avgtxpower) then delete it
		if (z%2 != 0):
			ws.delete_rows(z+1, 1)
		#if on an even index (i.e. we are on evm row) then keep it and append the avgtxpower onto it
		else:
			ws[(indexNumToColStr(keyParametersTX.index('NumValue'))) + (str(z+1))] = avgTXPowerColList[int(z/2)]
	
	#save the file!	
	wb.save(newSaveFileName)	
		
	return 1	


##################combines two files for non-similar test conditions####################
#########################NOT FINISHED###################
def combineNSParsedFile(file1, file2, combinedFileName):
	#load the file
	wb1 = load_workbook(file1)
	ws1 = wb1.active
	
	#load the file
	wb2 = load_workbook(file2)
	ws2 = wb2.active
	
	#load the test conditoins into memory for the 1st file 
#	testCondsTup1 = ws1[('B') + ':' + ('D')]
	testConds1 = ws1[('B') + ':' + ('E')]
	
	#load the test conditoins into memory for the 2nd file 
	testConds2 = ws2[('B') + ':' + ('E')]

	
	#add two new columns after the last ones to add more data
	ws1.insert_cols( len(keyParametersTX) + 2,2)
	
	#loop through ws2 and ws1 and see if there's matching test conditons. if so, add them in to the right of existing data. if not, add them in a new row
	for irow in range(len(testConds1[0]) - 1, 0, -1):
		for icol in range(len(testConds1)):
			ws1Value = testConds1[icol][irow].value
			ws2Value = testConds2[icol][irow].value
			#if all the test conditions including the power level match, then add in the new data to the right of the existing data
			if (ws1Value == ws2Value) & (icol == (len(testConds1) - 1)):
				ws1['H' + str(irow)] = ws2Value
	
	
	#loop through the saved data and index where there's a change in test conditions	
	# newDataIndexList = [0]
	# for irow in range(1, len(testConds1[0])):
		# for icol in range(len(testConds1)):
			# #print (testConds1[icol][irow].value)
			# prevVal = testConds1[icol][irow-1].value
			# if ((testConds1[icol][irow].value != prevVal) & (newDataIndexList[len(newDataIndexList) - 1] != irow)):
				# newDataIndexList.append(irow)
			
			
	#do the above... but BACKWARDS!
	#loop through the saved data and index where there's a change in test conditions	
	newDataIndexList = [0]
	for irow in range(len(testConds1[0]) - 1, 0, -1):
		for icol in range(len(testConds1)):
			#print (testConds1[icol][irow].value)
			prevVal = testConds1[icol][irow-1].value
			if ((testConds1[icol][irow].value != prevVal) & (newDataIndexList[len(newDataIndexList) - 1] != irow)):
				newDataIndexList.append(irow)			
			
	newDataIndexList.append(0)		
	newDataIndexList.pop(0)		
	
	print (newDataIndexList)
	
	#save the file!	
	wb1.save(combinedFileName)	
	
	
	
##################combines two files for SAME test conditions####################
def combineSParsedFileTX(file1, file2, combinedFileName):
	#load the file
	wb1 = load_workbook(file1)
	ws1 = wb1.active
	
	#load the file
	wb2 = load_workbook(file2)
	ws2 = wb2.active
	
	#add two new columns after the last ones to add more data
	ws1.insert_cols( len(keyParametersTX) + 2,2)
	
	#load up the 2nd parsed file's power and EVM
	mPowerCol2 = ws2['F']
	mEVMCol2 = ws2['G']
	
	#add the data from the 2nd workbook to the first
	for x in range(len(mPowerCol2)):
		ws1['H' + (str(x+1))] = mPowerCol2[x].value
		ws1['I' + (str(x+1))] = mEVMCol2[x].value
		
		
	#add in the row explaining what each column is
	ws1.insert_rows(1)
	for i in range(len(keyParametersTXHeader)):
		ws1[indexNumToColStr(i) + '1'] = keyParametersTXHeader[i]
		
	# #save the file!	
	wb1.save(combinedFileName)	

##################adds all the charts yo####################
def addChartsParsedTX(fileName, outputFileName):	
	#then load it up to work with
	wbc = load_workbook(fileName)
	wsc = wbc.active
	
	tPowerCol = wsc['E']
	
	#loop through the saved data and index where there's a change in power level	
	newPowerIndexList = []
	newPowerIndexList.append(2)
	
	#for irow in range(1, len(tPowerCol)):
	for irow in range(2, len(tPowerCol)):
		prevVal = tPowerCol[irow-1].value
		if (prevVal > tPowerCol[irow].value):
			newPowerIndexList.append(irow+1)
			
	for i in range(len(newPowerIndexList)):
		chart = ScatterChart()
		chart.style = 12
		chart.height = 3
		
		minxrow = newPowerIndexList[i]
		minyrow = newPowerIndexList[i]
		
		if (i == (len(newPowerIndexList) - 1)):
			maxrow = len(tPowerCol)
		else:
			maxrow = newPowerIndexList[i+1] - 1

		xValues1 = Reference(wsc, min_col=6, min_row=minxrow, max_row=maxrow)
		xValues2 = Reference(wsc, min_col=8, min_row=minxrow, max_row=maxrow)
	
		yValues1 = Reference(wsc, min_col=7, min_row=minyrow, max_row=maxrow)
		yValues2 = Reference(wsc, min_col=9, min_row=minyrow, max_row=maxrow)
		series1 = Series(yValues1, xValues1, title='b1 evm')
		series2 = Series(yValues2, xValues2, title='b2 evm')
		
		chart.series.append(series1)
		chart.series.append(series2)
		wsc.add_chart(chart, ("J"+str(newPowerIndexList[i])))
		
	
	#save file
	wbc.save(outputFileName)
	
	

#parseFileTX(INPUTFILE1, OUTPUTFILE1)
#parseFileTX(INPUTFILE2, OUTPUTFILE2)
	
combineSParsedFileTX(OUTPUTFILE1, OUTPUTFILE2, 'test.xlsx')	
	
addChartsParsedTX('test.xlsx', 'scatter.xlsx')
